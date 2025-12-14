#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys


chrome_path = "C:\seleniumwepscraper\chromedriver.exe"
driver = webdriver.Chrome(chrome_path)
driver.implicitly_wait(15)#รอจนกว่าจะเปิดเสร็จ ออกก่อนเวลา
driver.maximize_window()
driver.get("https://landsmaps.dol.go.th/")

driver.implicitly_wait(10)
try: 
    esc_button = driver.find_element_by_css_selector('body.modal-open:nth-child(2) div.modal.fade.show div.modal-dialog.modal-lg.Clang div.modal-content div:nth-child(3) > button.btn.btn-danger')
    esc_button.click()
    esc_button = driver.find_element_by_css_selector('body.modal-open:nth-child(2) div.modal.fade.show div.modal-dialog.modal-lg.Clang div.modal-content div:nth-child(3) > button.btn.btn-danger')
    esc_button.click()
except:
    print('No element with this class name. SKipping 1 ')
    
driver.implicitly_wait(10)    
try: 
    esc_button = driver.find_element_by_css_selector('body.modal-open:nth-child(2) div.modal.fade.show div.modal-dialog.modal-lg.Clang div.modal-content div.modal-header:nth-child(1) button.close > i.fas.fa-times-circle')
    esc_button.click()
except:
    print('No element with this class name. SKipping 2 ')
    
login_button = driver.find_element_by_css_selector('#btnLoginMenu1')
login_button.click()

login_button = driver.find_element_by_css_selector('#lmw02_placeholder01')
#login_button.clear()
login_button.send_keys("@hotmail.com")
login_button = driver.find_element_by_css_selector('#lmw02_placeholder02')
#login_button.clear()
login_button.send_keys("password")

driver.implicitly_wait(2) 
enter_button = driver.find_element_by_css_selector('body.modal-open:nth-child(2) div.modal.fade.show:nth-child(2) div.modal-dialog.modal-lg div.modal-content div.modal-body.__background div.tab-content:nth-child(3) div.tab-pane.fade.show.active:nth-child(1) form.needs-validation05:nth-child(2) div.row.___margin:nth-child(4) div.col-12.col-md-12.col-sm-12 > button.btn.btn-primary:nth-child(2)')
enter_button.click()


# In[30]:


import time


wb = load_workbook("C:\\seleniumwepscraper\\testwork06.xlsx")# path เบิ้ล '\'
print("****************ได้ excel แล้ว***************")
ws = wb.worksheets[0]
ws.cell(1,16).value= "1"


for row in ws.iter_rows():
    
    if ws.cell( row[0].row, 8 ).value != None : #เช็คช่องว่าง
        #print(ws.cell( row[0].row, 1 ).value)
        print(str(ws.cell( row[0].row, 1 ).value)+"  " + ws.cell( row[0].row, 8 ).value )
        duplicate = ws.cell( row[0].row, 8 ).value
        continue
        
    collect_box = driver.find_elements_by_css_selector("div.col-8.text-sm-left")
    
   
    #start 5
    ws.cell( row[0].row, 5 ).value = collect_box[6].text #จังหวัด
    ws.cell( row[0].row, 6 ).value = collect_box[5].text #อำเภอ
    ws.cell( row[0].row, 7 ).value = collect_box[4].text #ตำบล
    ws.cell( row[0].row, 8 ).value = collect_box[0].text #เลขโฉนด
    ws.cell( row[0].row, 9 ).value = collect_box[2].text #เลขที่ดิน
    ws.cell( row[0].row, 16 ).value = collect_box[1].text #หน้าสำรวจ
    ws.cell( row[0].row, 17 ).value = collect_box[3].text #เลขระวาง
    
    clean7 = collect_box[7].text.split() #แยกไร่-งาน-ตรวา
    
    clean8 = collect_box[8].text.split() #ค่ารางวัด ต่อ ตรวา
    
    try:#บางที่ ไม่มีลูกน้ำ น้อยกว่า1000
        clean8 = clean8[0].replace(',', '')
    except:
        clean8 = collect_box[8].text.split()
        clean8 = clean8[0]
        
    clean9 = collect_box[9].text.split(',') #แยกGPS
    
    ws.cell( row[0].row, 10 ).value = clean7[0]
    ws.cell( row[0].row, 11 ).value = clean7[2]
    ws.cell( row[0].row, 12 ).value = clean7[4]
    ws.cell( row[0].row, 13 ).value = clean8
    ws.cell( row[0].row, 14 ).value = clean9[0]
    ws.cell( row[0].row, 15 ).value = clean9[1]
    
    #duplicate = 1 #ใช้ตอนเริ่มโปรแกรมที่ไม่มีตัวก่อนหน้า
    if  duplicate == ws.cell( row[0].row, 8 ).value : #กันกดซ้ำตัว"ก่อนหน้า"
        print(str(ws.cell( row[0].row, 1 ).value)+"  "+ws.cell( row[1].row, 8 ).value +" ไม่เซฟ")
        break
        
    print(str(ws.cell( row[0].row, 1 ).value)+"  " + ws.cell( row[0].row, 8 ).value +" เซฟ")
    wb.save("C:\\seleniumwepscraper\\testwork06.xlsx")
    break
    time.sleep(3)


# In[ ]:




