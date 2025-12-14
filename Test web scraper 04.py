#!/usr/bin/env python
# coding: utf-8

# In[9]:


from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys


chrome_path = "C:\seleniumwepscraper\chromedriver.exe"
driver = webdriver.Chrome(chrome_path)
driver.implicitly_wait(15)#รอจนกว่าจะเปิดเสร็จ ออกก่อนเวลา
driver.maximize_window()
driver.get("https://landsmaps.dol.go.th/")

driver.implicitly_wait(10)
try: 
    esc_button = driver.find_element_by_css_selector('body.modal-open:nth-child(2) div.modal.fade.show div.modal-dialog.modal-lg.Clang div.modal-content div:nth-child(3) > button.btn.btn-danger')
    esc_button.click()
    esc_button = driver.find_element_by_css_selector('body.modal-open:nth-child(2) div.modal.fade.show div.modal-dialog.modal-lg.Clang div.modal-content div:nth-child(3) > button.btn.btn-danger')
    esc_button.click()
except:
    print('No element with this class name. SKipping 1 ')
    
driver.implicitly_wait(10)    
try: 
    esc_button = driver.find_element_by_css_selector('body.modal-open:nth-child(2) div.modal.fade.show div.modal-dialog.modal-lg.Clang div.modal-content div.modal-header:nth-child(1) button.close > i.fas.fa-times-circle')
    esc_button.click()
except:
    print('No element with this class name. SKipping 2 ')
    
login_button = driver.find_element_by_css_selector('#btnLoginMenu1')
login_button.click()

login_button = driver.find_element_by_css_selector('#lmw02_placeholder01')
#login_button.clear()
login_button.send_keys("@hotmail.com")
login_button = driver.find_element_by_css_selector('#lmw02_placeholder02')
#login_button.clear()
login_button.send_keys("passowrd")

driver.implicitly_wait(2) 
enter_button = driver.find_element_by_css_selector('body.modal-open:nth-child(2) div.modal.fade.show:nth-child(2) div.modal-dialog.modal-lg div.modal-content div.modal-body.__background div.tab-content:nth-child(3) div.tab-pane.fade.show.active:nth-child(1) form.needs-validation05:nth-child(2) div.row.___margin:nth-child(4) div.col-12.col-md-12.col-sm-12 > button.btn.btn-primary:nth-child(2)')
enter_button.click()


# In[10]:


import time
#เลข XPATH

number_province_tranform = { 'พระนครศรีอยุธยา' : 34 ,'ฉะเชิงเทรา' : 9 , 'สมุทรปราการ' : 61 , 'ชุมพร' : 13 , 
                            'เชียงราย' : 14 
                           
                          
}

number_aumpher_tranform = {'เมืองกระบี่' : 1 , 'เขาพนม' : 2 , 'เกาะลันตา' : 3 ,
                           'คลองท่อม' : 4 , 'อ่าวลึก' : 5 , 'ปลายพระยา' : 6 ,
                           'ลำทับ' : 7 , 'เหนือคลอง' : 8 ,

                           'พระนคร' : 1 , 'ดุสิต' : 2 , 'หนองจอก' :  3,
                           'บางรัก' : 4,  'บางเขน' : 5 , 'บางกะปิ' : 6 ,
                           'ปทุมวัน' : 7 , 'ป้อมปราบศัตรูพ่าย' : 8 , 'พระโขนง' : 9 ,
                           'มีนบุรี' : 10 , 'ลาดกระบัง' : 11 , 'ยานนาวา' : 12 ,
                           'สัมพันธวงศ์' : 13 , 'พญาไท' : 14 , 'ธนบุรี' : 15 ,
                           'บางกอกใหญ่' : 16 , 'ห้วยขวาง' : 17 , 'คลองสาน' : 18 ,
                           'ตลิ่งชัน' : 19 , 'บางกอกน้อย' : 20 , 'บางขุนเทียน' : 21 ,
                           'ภาษีเจริญ' : 22 , 'หนองแขม' : 23 , 'ราษฎร์บูรณะ' : 24 ,
                           'บางพลัด' : 25 , 'ดินแดง' : 26 , 'บึงกุ่ม' : 27 ,
                           'สาทร' : 28 , 'บางซื่อ' : 29 , 'จตุจักร' : 30 ,
                           'บางคอแหลม' : 31 , 'ประเวศ' : 32 , 'คลองเตย' : 33 ,
                           'สวนหลวง' : 34 , 'จอมทอง' : 35 , 'ดอนเมือง' : 36 ,
                           'ราชเทวี' : 37 , 'ลาดพร้าว' : 38 , 'วัฒนา' : 39 ,
                           'บางแค' : 40 , 'หลักสี่' : 41 , 'สายไหม' : 42 ,
                           'คันนายาว' : 43 , 'สะพานสูง' : 44 , 'วังทองหลาง' : 45 ,
                           'คลองสามวา' : 46 , 'บางนา' : 47 , 'ทวีวัฒนา' : 48 ,
                           'ทุ่งครุ' : 49 , 'บางบอน' : 50 ,
                          
                           'พระนครศรีอยุธยา' : 1 , 'ท่าเรือ' : 2 , 'นครหลวง' : 3 ,
                           'บางไทร' : 4 , 'บางบาล' : 5 , 'บางปะอิน' : 6 ,
                           'บางปะหัน' : 7 , 'ผักไห่' : 8 , 'ภาชี' : 9 ,
                           'ลาดบัวหลวง' : 10 , 'วังน้อย' : 11 , 'เสนา' : 12 ,
                           'บางซ้าย' : 13 , 'อุทัย' : 14 , 'มหาราช' : 15 ,
                           'บ้านแพรก' : 16 ,
                          
                           'แปลงยาว' : 9 ,
                          
                           'บางบ่อ' : 2 ,
                          
                           'เมืองชุมพร' : 1 ,
                          
                           'เทิง' : 4 
                          
} 

wb = load_workbook("C:\\seleniumwepscraper\\test.xlsm")# path เบิ้ล '\'
print("****************ได้ excel แล้ว***************")
print("****************ได้ excel แล้ว***************")
ws = wb.worksheets[0]
ws.cell(1,16).value= "1"

for row in ws.iter_rows():
   num_Province = number_province_tranform[str(row[0].value)]
   num_Aumpher = number_aumpher_tranform[str(row[1].value)] +1
   num_titleland = int(row[2].value)
   
   print(number_province_tranform[str(row[0].value)] )
   print(number_aumpher_tranform[str(row[1].value)] )
                                  
   driver.implicitly_wait(3)
   
   xpath_province = "//body/nav[1]/form[1]/div[1]/select[1]/option[" +str(num_Province)+"]"
   fill_province = driver.find_element_by_xpath(xpath_province)
   fill_province.click()
   
   xpath_aumpher = "//body/nav[1]/form[2]/div[1]/select[1]/option[" +str(num_Aumpher)+"]"
   fill_aumpher = driver.find_element_by_xpath(xpath_aumpher)
   fill_aumpher.click()
   
   fill_titleland = driver.find_element_by_xpath("//input[@id='txtparcelno']")
   fill_titleland.clear()
   fill_titleland.send_keys(num_titleland)
   
   time.sleep(2)
   search = driver.find_element_by_xpath("//button[@id='btnSearch']")
   search.click()
   
   time.sleep(3)
   driver.implicitly_wait(5)
   collect_box = driver.find_elements_by_css_selector("div.col-8.text-sm-left")
   
  
   #start 5
   ws.cell( row[0].row, 5 ).value = collect_box[6].text #จังหวัด
   ws.cell( row[0].row, 6 ).value = collect_box[5].text #อำเภอ
   ws.cell( row[0].row, 7 ).value = collect_box[4].text #ตำบล
   ws.cell( row[0].row, 8 ).value = collect_box[0].text #เลขโฉนด
   ws.cell( row[0].row, 9 ).value = collect_box[2].text #เลขที่ดิน
   ws.cell( row[0].row, 10 ).value = collect_box[1].text #หน้าสำรวจ
   ws.cell( row[0].row, 11 ).value = collect_box[3].text #เลขระวาง
   
   clean7 = collect_box[7].text.split() #แยกไร่-งาน-ตรวา
   
   clean8 = collect_box[8].text.split() #ค่ารางวัด ต่อ ตรวา
   
   try:#บางที่ ไม่มีลูกน้ำ น้อยกว่า1000
       clean8 = clean8[0].replace(',', '')
   except:
       clean8 = collect_box[8].text.split()
       clean8 = clean8[0]
       
   clean9 = collect_box[9].text.split(',') #แยกGPS
   
   ws.cell( row[0].row, 12 ).value = clean7[0]
   ws.cell( row[0].row, 13 ).value = clean7[2]
   ws.cell( row[0].row, 14 ).value = clean7[4]
   ws.cell( row[0].row, 15 ).value = clean8
   ws.cell( row[0].row, 16 ).value = clean9[0]
   ws.cell( row[0].row, 17 ).value = clean9[1]
   
   wb.save("C:\\seleniumwepscraper\\result_test.xlsx")
   time.sleep(3)


# In[ ]:




